from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Alignment


def copy_gs(sheet, source, target):
    '''
    自定义单元格格式复制函数
    :param sheet: sheet
    :param source: 根据
    :param target: 调整，如"A1"
    :return:
    '''

    source_cell = sheet[source]  # 源,根据
    target_cell = sheet[target]  # 目标

    target_cell.data_type = source_cell.data_type
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)


'''解除出<实测值>合并单元格，重新合并居中'''


def scz(sheet, s_position, d_position, cha_num):
    i = 1
    while True:
        try:
            sheet.unmerge_cells(start_row=2, end_row=2,
                                start_column=s_position, end_column=d_position - i)
            break
        except Exception as e:
            i += 1
            if i > 50:
                print("<实测>值未合并剧中或仅占一个单元格")
                break
            print("执行", i - 1)

    # 避免接触合并剧中避免报错
    sheet.merge_cells(start_row=2, end_row=2,
                      start_column=s_position, end_column=s_position + cha_num)
    align = Alignment(horizontal='center', vertical='center')
    sheet.cell(2, s_position).alignment = align


'''比较实测值与预期值是否匹配'''


def scz_yqz(sheet, start_position, end_position):
    bj = 0
    for i in range(end_position - start_position):
        start_char, end_char = chr(ord("A") + start_position + i - 1), chr(
            ord("A") + end_position + i - 1)  # 转换为大写字母，列值
        if sheet["{}4".format(end_char)].value != sheet["{}4".format(start_char)].value:
            print('在{0}中单元格{1}4与单元格{2}的值不匹配'.format(sheet.title, end_char,start_char))
            bj = 1
    return bj

'''实测值赋值及自动换行调整'''
def scz_value(sheet, start_position, end_position):
    for i in range(end_position - start_position):
        start_char, end_char = chr(ord("A") + start_position + i - 1), chr(
            ord("A") + end_position + i - 1)  # 转换为大写字母，列值
        # print("在{}根据{}列调整{}列".format(sheet, start_char, end_char))
        # sheet.column_dimensions[end_char].width = sheet.column_dimensions["{}1".format(start_char)].width
        # sheet.column_dimensions[end_char].width = sheet.column_dimensions[start_char].width
        sheet["{}4".format(end_char)].value = sheet["{}4".format(start_char)].value
        for j in range(6, sheet.max_row + 1):  # 从6行遍历至最大行，调整实测值单元格格式，并自动换行
            source = "{0}6".format(chr(ord("A") + end_position - 1))  # 根据6行实测值第一列为根据，调整格式
            target = "{0}{1}".format(end_char, j)
            align = Alignment(wrapText=True)  # 自动换行
            sheet.cell(6, end_position).alignment = align  # 6行实测值第一列，格式添加自动换行
            copy_gs(sheet, source, target)


def main():
    wb = load_workbook("example3.xlsx")
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)
    # insert_rows(3)
    for i in range(num_sheets):
        if 0 < i < num_sheets - 1:
            sheet = wb[sheet_names[i]]  # 选择指定的sheet
            print(sheet)
            cell_positions = []
            # 遍历第二行的所有单元格
            for c in sheet.iter_rows(min_row=2, max_row=2):
                for cell in c:
                    # print(cell)
                    if cell.value == "实测值":
                        cell_positions.append([cell.row, cell.column])
                    if cell.value == "预期值":
                        cell_positions.append([cell.row, cell.column])
                    if cell.value == "单项测试结论（P/Fail）":
                        cell_positions.append([cell.row, cell.column])
                        # print("you")
            print(cell_positions)
            cha = 2 * cell_positions[1][1] - cell_positions[0][1] - cell_positions[2][1]  # 实测值的列数-预期值的列数
            '''函数输入值'''
            start_position, end_position = cell_positions[0][1], cell_positions[1][1]
            s_position, d_position = cell_positions[1][1], cell_positions[2][1]
            cha_num = cell_positions[1][1] - cell_positions[0][1] - 1

            bj = scz_yqz(sheet, start_position, end_position)
            if cha != 0 or bj:
                # 首先解除掉<单项测试结论（P/Fail）>的单元格合并
                sheet.unmerge_cells(range_string='{0}2:{0}5'.format(chr(ord("A") + cell_positions[2][1] - 1)))
                print("在文件**{}中<实测值>与<预期值>不匹配，开始自动添加实测值，并调整格式".format(sheet_names[i]))
                # 插入实测值的列数-预期值的列数的cha个列
                for s in range(cell_positions[2][1], cell_positions[2][1] + cha):
                    sheet.insert_cols(cell_positions[2][1])

                # 对插入后新位置的<单项测试结论（P/Fail）>的单元格合并
                sheet.merge_cells(range_string='{0}2:{0}5'.format(chr(ord("A") + cell_positions[2][1] + cha - 1)))
                # 为实测值新增单元格进行格式复制
                for x in range(cell_positions[1][1] - 1, cell_positions[2][1] + cha - 1):
                    copy_gs(sheet, "E4", "{}4".format(chr(ord("A") + x)))  # E4单元格的格式复制给实测值的第四行
                    st = "{0}5".format(chr(ord("A") + cell_positions[1][1] - 1))
                    copy_gs(sheet, st, "{0}3".format(chr(ord("A") + x)))  # st单元格的格式复制给实测值的第3行
                    copy_gs(sheet, st, "{0}5".format(chr(ord("A") + x)))  # st单元格的格式复制给实测值的第5行,st为实测值的第一个列第三行单元格
                    print("***", chr(ord("A") + cell_positions[1][1] - 1), st)
                    # copy_gs("{0}5", "{1}5".format(chr(ord("A") + cell_positions[1][1] - 1), chr(ord("A") + x)))
                    print("E4", "{}4".format(chr(ord("A") + x)))

                scz(sheet, s_position, d_position, cha_num)
                scz_value(sheet, start_position, end_position)
            else:
                print("<实测值>与<预期值>匹配，仅并调整格式")
                print("在文件**{}中<实测值>与<预期值>匹配，仅对<实测值>区域格式调整".format(sheet_names[i]))
                scz(sheet, s_position, d_position, cha_num)
                scz_value(sheet, start_position, end_position)

    wb.save("example4.xlsx")


if __name__ == '__main__':
    main()
