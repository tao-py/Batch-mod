import openpyxl
from tkinter import *
from tkinter import filedialog

'''定义全局变量'''
st = []
st_cut = []
jilu = []
# 创建一个函数用于选择文件
def select_files():
    # 获取所选文件的路径
    global file_path
    file_path = filedialog.askopenfilenames(initialdir='/path/to/folder', filetypes=[('Excel files', '*.xlsx')])
    file_entry.delete(0, END)
    file_entry.insert(0, file_path)
    print("选择的文件有：")
    for filename in file_path:
        print("{}".format(filename))
    return file_path

'''自定义设置窗口在屏幕的正中间的函数'''
def centre_window(ww,wh):
    sw = root.winfo_screenwidth()#获取屏幕的宽度
    sh = root.winfo_screenheight()#获取屏幕的高度
    #设置窗口宽度和高度;ww,wh=600,320
    x,y=(sw-ww)/2,(sh-wh)/2
    cen="%dx%d+%d+%d"%(ww,wh,x,y)
    return cen

def get_original_text():
    text = original_text.get("1.0", "end-1c")
    print("Original Text:", text)
    return text
def get_replaced_text():
    text = replaced_text.get("1.0", "end-1c")
    print("Replaced Text:", text)
    return text

def sure_txt():
    O_txt = get_original_text()
    R_txt = get_replaced_text()
    global st
    '''全局变量至关重要'''
    st.append([O_txt, R_txt])
    return st

# 创建替换函数
def replace_excel(filename):
    global jilu
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                # print(cell.value)
                for i in range(len(st_cut)):
                    # print(st_cut[i][0],st_cut[i][1])
                    if str(cell.value) == st_cut[i][0]:
                        jilu.append([filename,sheet_name,cell.value,st_cut[i][1]])
                        '''STR()避免了数字替换不了的情况，但替换后数字格式为文本'''
                        # print("xiangdeng")
                        cell.value = st_cut[i][1]
    wb.save(filename)

'''遍历所选取文件的函数并执行替换函数'''
def bianli_files():
    for filename in file_path:
        replace_excel(filename)

###最后一步###
def next_step():
    global st
    '''删除重复输入'''
    global st_cut
    for i in range(len(st)):
        if i == len(st) - 1:
            st_cut.append(st[i])
        else:
            if st[i] == st[i + 1]:
                continue
            else:
                st_cut.append(st[i])
    print("待替换的值有：")
    for i in range(len(st_cut)):
        print("{0} 替换为：{1}".format(st_cut[i][0],st_cut[i][1]))
    run_function()
    jilu_value()
    pass
def run_task():

    popup_window = Toplevel(root)
    popup_window.title("任务执行完成")
    popup_window.geometry(centre_window(200,150))
    popup_label = Label(popup_window, text="替换完成！")
    popup_label.pack(pady=20)
    popup_button = Button(popup_window, text="关闭", command=popup_window.destroy)
    popup_button.pack(pady=10)

'''为选择文件执行报错异常处理，弹窗'''
def run_function():
    try:
        bianli_files()
    except Exception as e:
        popup_window = Toplevel(root)
        popup_window.title("请选择EXCEL文件")
        popup_window.geometry(centre_window(200, 150))
        popup_label = Label(popup_window, text="请先选择文件！")
        popup_label.pack(pady=20)
        popup_button = Button(popup_window, text="关闭", command=popup_window.destroy)
        popup_button.pack(pady=10)
    else:
        run_task()

def jilu_value():
    if jilu == []:
        print("{:*^30}".format("未执行成功"))
        print("{:*^30}".format("检查替换值或文件是否存在"))
    else:
        print("{:*^30}".format("执行结束"))
        print("{:*^30}".format("点击<开始替换>可继续执行"))
        fl = jilu[0][0]
        a = fl.split("/")
        fl = "/".join(a[:-1])
        with open('{}/替换记录.txt'.format(fl), 'a') as f:
            for l in jilu:
                f.write("在{0}文件Excel表{1}中，已将{2}替换为{3}\n".format(*l))

# 创建GUI界面
root = Tk()
root.title("Excel批量替换")
root.geometry(centre_window(600,320))
# 创建按钮
# run_button = Button(root, text="执行任务", command=run_task)
# run_button.pack(pady=20)

file_label = Label(root, text="Excel文件：")
file_label.grid(row=0, column=0)

file_entry = Entry(root,width=50)
file_entry.grid(row=0, column=1)

file_button = Button(root, text="选择文件", command=select_files)
file_button.grid(row=0, column=2)

###
original_text_label = Label(root, text="替换前内容:")
original_text_label.grid(row=2, column=1)
original_text = Text(root, height=5, width=50)
original_text.grid(row=3, column=1)

replaced_text_label = Label(root, text="替换后内容:")
replaced_text_label.grid(row=4, column=1)
replaced_text = Text(root, height=5, width=50)
replaced_text.grid(row=5, column=1)

get_replaced_text_button = Button(root, text="确定", command=sure_txt)
get_replaced_text_button.grid(row=5, column=2)

next_button = Button(root, text="开始替换", command= next_step)
next_button.grid(row=6, column=1)

root.mainloop()