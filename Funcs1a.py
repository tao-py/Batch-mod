from tkinter import filedialog
import openpyxl
from tkinter import *
import datetime
import win32com.client as win32

'''定义全局变量'''
st_cut = []
st = []
jilu = []

def select_files():
    """
    获取所选文件的路径
    :return: file_path
    """
    global file_path
    file_path=()
    file_path = filedialog.askopenfilenames(initialdir='/path/to/folder', filetypes=[('Excel files', '*.xlsx')])
    print("选择的文件有：")
    for filename in file_path:
        print("{}".format(filename))
    return file_path

def centre_window(ww,wh,root):
    """
    自定义设置窗口在屏幕的正中间的函数
    :param ww: 设置高度参数
    :param wh: 设置宽度参数
    :param root: 窗口赋予
    :return: cen用于geometry()内
    """
    sw = root.winfo_screenwidth()#获取屏幕的宽度
    sh = root.winfo_screenheight()#获取屏幕的高度
    x,y = (sw-ww)/2,(sh-wh)/2
    cen = "%dx%d+%d+%d"%(ww,wh,x,y)
    return cen
# 创建替换函数
class GUI_text :
    def __init__(self, window):
        self.window = window
        global original_text, replaced_text
        original_text = Text(self.window, height=3, width=50)
        original_text.grid(row=3, column=1)
        replaced_text = Text(self.window, height=3, width=50)
        replaced_text.grid(row=5, column=1)
    def get_text(self):
        O_txt = original_text.get("1.0", "end-1c")
        print("Original Text:", O_txt)
        R_txt = replaced_text.get("1.0", "end-1c")
        print("Replaced Text:", R_txt)
        '''全局变量至关重要'''
        global st
        st.append([O_txt, R_txt])
        return st

def replace_excel(filename):
    global jilu
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                for i in range(len(st_cut)):
                    if str(cell.value) == st_cut[i][0]:
                        jilu.append([filename, sheet_name, cell.value, st_cut[i][1]])
                        '''STR()避免了数字替换不了的情况，但替换后数字格式为文本'''
                        cell.value = st_cut[i][1]
    wb.save(filename)


def run_function():
    try:
        for filename in file_path:
            replace_excel(filename)
    except Exception as e:
        popup_window = Toplevel()
        popup_window.title("请选择EXCEL文件")
        popup_window.geometry(centre_window(200, 150,popup_window))
        popup_label = Label(popup_window, text="请先选择文件！")
        popup_label.pack(pady=20)
        popup_button = Button(popup_window, text="关闭", command=popup_window.destroy)
        popup_button.pack(pady=10)
    else:
        run_task()

def run_task():
    popup_window = Toplevel()
    popup_window.title("任务执行完成")
    popup_window.geometry(centre_window(200, 150, popup_window))
    popup_label = Label(popup_window, text="替换完成！")
    popup_label.pack(pady=20)
    popup_button = Button(popup_window, text="关闭", command=popup_window.destroy)
    popup_button.pack(pady=10)

def jilu_value():
    if jilu == []:
        print("{:*^30}".format("未执行成功"))
        print("{:*^30}".format("检查替换值或文件是否存在"))
    else:
        print("{:*^30}".format("执行结束"))
        print("{:*^30}".format("点击<开始替换>可继续执行"))
        fl = jilu[0][0]
        a = fl.split("/")
        fl = "/".join(a[:-1])
        with open('{}/替换记录.txt'.format(fl), 'w') as f:
            # 获取当前时间
            now = datetime.datetime.now()
            time_str = now.strftime("%Y-%m-%d %H:%M")
            f.write("执行时间："+time_str + "\n")
            for l in jilu:
                f.write("在{0}文件Excel表{1}中，已将{2}替换为{3}\n".format(*l))

def begin_run():
    for i in range(len(st)):
        if i == len(st) - 1:
            st_cut.append(st[i])
        else:
            if st[i] == st[i + 1]:
                continue
            else:
                st_cut.append(st[i])
    print("待替换的值有：")
    for i in range(len(st_cut)):
        print("{0} 替换为：{1}".format(st_cut[i][0], st_cut[i][1]))
    run_function()
    jilu_value()
    pass
# [O_tab,R_tab]
def OR_must(filename):
    wb = openpyxl.load_workbook(filename)
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)
    global OR_tab
    OR_tab = []
    for i in range(num_sheets):
        if 0 < i < num_sheets - 1:
            sheet = wb[sheet_names[i]]  # 选择指定的sheet
            cell_positions = []
            # 遍历工作表中的所有单元格
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == "实测值":
                        cell_positions.append([i,cell.row, cell.column])
                    if cell.value == "预期值":
                        cell_positions.append([i,cell.row, cell.column])
            # 输出符合条件的单元格的行列位置
            OR_tab.append(cell_positions)

def shicezhi(filename,OR_tab):
    # 创建 Excel 应用程序对象
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    # 循环遍历所有 Excel 文件
    workbook = excel.Workbooks.Open(filename)
    a = filename.split("/")
    print("----------->在文件:{}中".format(a[-1]))
    for i in range(len(OR_tab)):
        worksheet = workbook.Worksheets(i+2)
        star=OR_tab[i][0][2]
        over=OR_tab[i][1][2]
        for j in range(over-star):
            a_col_width = worksheet.Columns(star+j).ColumnWidth
            # 将 over+j列的宽度设置为与star+j列相同
            worksheet.Columns(over+j).ColumnWidth = a_col_width
            print("***Case-{0}中根据第{1}列--调整Case-{2}中调整第{3}列".format(i + 1, star+j , i + 1, over+j))
            # 保存修改后的 Excel 文件
        workbook.Save()
    # 关闭 Excel 文件
    workbook.Close()
    # 退出 Excel 应用程序
    excel.Quit()



def run_function1():
    try:
        for filename in file_path:
            OR_must(filename)
            shicezhi(filename, OR_tab)
    except Exception as e:
        popup_window = Toplevel()
        popup_window.title("请选择EXCEL文件")
        popup_window.geometry(centre_window(200, 150,popup_window))
        popup_label = Label(popup_window, text="请先选择文件！")
        popup_label.pack(pady=20)
        popup_button = Button(popup_window, text="关闭", command=popup_window.destroy)
        popup_button.pack(pady=10)
    else:
        print("{:*^30}".format("执行结束"))
        print("{:*^30}".format("返回可继续执行"))
        run_task()




