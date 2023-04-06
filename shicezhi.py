import tkinter as tk
from tkinter import filedialog
import openpyxl

root = tk.Tk()
root.withdraw()
fname = filedialog.askopenfile()
print(fname)


'''自定义文件名剪切函数'''
def flcut (filename):
    st = str(filename)
    begin = st.index("'")
    over = st.index("'", begin + 1)
    st = st[begin+1:over]
    return st

'''自定义--函数'''
def odd_omm (sheet):
    # 遍历指定范围内的单元格并检查它们的值
    for row in sheet.iter_rows(min_row=6, max_row=8, min_col=2, max_col=2):
        for cell in row:
            print(cell.value)
            if "OMM" == str(cell.value):
                print("有OMM")
                return "OMM"
                break
            elif "ODD" == str(cell.value):
                print("有ODD")
                return "ODD"
                break
            else:
                continue
    return 0

'''自定义--函数'''
def rename(wb):
    sheet_names = wb.sheetnames
    num_sheets = len(sheet_names)
    # 修改sheet名
    for i in range(num_sheets):
        if 0 < i < num_sheets-1:
            sheet = wb[sheet_names[i]]
            sheet.title = 'sheet-{}'.format(i)
    for i in range(num_sheets):
        if 0 < i < num_sheets-1:
            sheet = wb['sheet-{}'.format(i)]
            sheet.title = 'Case-{}'.format(i)

# worksheet = wb.active
# print(wb.active)

'''自定义--函数'''
def movesheet (sheet,wb):
    # 获取指定名称的工作表对象
    # worksheet = wb.worksheets[wb.sheetnames.index(sheet)]
    sheet_index = wb.index(sheet)
    target_index = len(wb.worksheets) - 2
    print(sheet_index)
    print(len(wb.worksheets))
    wb.move_sheet(sheet, offset=target_index-sheet_index)
    #offset表示偏移量，如果偏移量为1，则表示将工作表移动到当前位置的下一个位置，如果偏移量为-1，则表示将工作表移动到当前位置的上一个位置。

# 打开Excel文件
wb = openpyxl.load_workbook(flcut(fname))
# 获取所有的sheet名
sheet_names = wb.sheetnames
num_sheets = len(sheet_names)  # 获取工作表数量
print(sheet_names)
jilu = 0
for i in range(num_sheets):
    if 0 < i < num_sheets-1:
        sheet = wb[sheet_names[i]]      # 选择指定的sheet
        if odd_omm(sheet):
            print("yes{}".format(odd_omm(sheet)))
            if odd_omm(sheet) == "OMM":
                print("yidong")
                movesheet(sheet, wb)
                jilu = 1
if jilu:
    rename(wb)

wb.save('example.xlsx')  # 保存Excel文件
